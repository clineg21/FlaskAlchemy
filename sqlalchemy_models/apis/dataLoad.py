from flask import jsonify
from flask_restx import Namespace, Resource
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app import sessions
from models.ClientModels import Clients


api = Namespace('dataLoad', 'Clients APIs')

@api.route("/")
class Data(Resource):
    def get(self):
        wb = load_workbook(filename='/Users/gracecline/Documents/FlaskAlchemy/sqlalchemy_models/FlaskAlchemy.xlsx')
        ws = wb.active

        my_list = []

        last_column = len(list(ws.columns))
        last_row = len(list(ws.rows))

        for row in range(2, last_row + 1):
            my_dict = {}
            for column in range(1, last_column + 1):
                column_letter = get_column_letter(column)
                if row > 1:
                    my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
            my_list.append(my_dict)
        return jsonify(my_list)

    def post(self):
        wb = load_workbook(filename='/Users/gracecline/Documents/FlaskAlchemy/sqlalchemy_models/FlaskAlchemy.xlsx')
        ws = wb.active

        my_list = []

        last_column = len(list(ws.columns))
        last_row = len(list(ws.rows))

        for row in range(2, last_row + 1):
            my_dict = {}
            for column in range(1, last_column + 1):
                column_letter = get_column_letter(column)
                if row > 1:
                    my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
            my_list.append(my_dict)
        for item in my_list:
            client1 = Client(first_name=item["First Name"],
                                last_name=item["Last Name"],
                                age=item["Age"],
                                education=item["Education"],
                                occupation=item["Occupation"],
                                experience=item["Experience"],
                                salary=item["Salary"],
                                num_children=["Number of Children"])
            session.add(client1)
            session.commit()
        return 'Successfully Put Data In Database'